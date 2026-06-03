"""
本地文件扫描器 - JobTracer Module a) 数字足迹扫描器
支持异步扫描本地文件，提取文本预览
"""

import asyncio
import hashlib
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import aiofiles
try:
    import fitz  # PyMuPDF - optional, for PDF text extraction
except ImportError:
    fitz = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ============================================================
# 常量配置
# ============================================================

# 支持的文件类型及扩展名
SUPPORTED_EXTENSIONS = {'.md', '.txt', '.pdf', '.doc', '.docx', '.xlsx', '.pptx', '.py', '.js', '.ts'}

# 排除的目录（不扫描）
EXCLUDE_DIRS = {'node_modules', '.git', '__pycache__', '.venv', 'venv', 'build', 'dist', '.idea', '.vscode', 'node_modules', 'target', '.pytest_cache', '.mypy_cache'}

# 敏感文件名关键词（跳过）
SENSITIVE_FILENAMES = {'password', 'key', 'token', 'secret', '.env', 'credentials', 'config.ini', 'secrets', '.git-credentials', 'oauth'}

# 文件大小限制（5MB）
MAX_FILE_SIZE = 5 * 1024 * 1024

# content_preview 最大字符数
PREVIEW_MAX_CHARS = 200

# 文件类型分类
FILE_TYPE_MAP = {
    '.md': 'document',
    '.txt': 'document',
    '.pdf': 'document',
    '.doc': 'document',
    '.docx': 'document',
    '.xlsx': 'spreadsheet',
    '.pptx': 'presentation',
    '.py': 'code',
    '.js': 'code',
    '.ts': 'code',
}


# ============================================================
# 辅助函数
# ============================================================

def is_sensitive_file(path: Path) -> bool:
    """检查文件名是否包含敏感关键词"""
    name_lower = path.name.lower()
    return any(sensitive in name_lower for sensitive in SENSITIVE_FILENAMES)


def should_exclude_dir(path: Path) -> bool:
    """检查路径是否在排除目录中"""
    parts = path.parts
    return any(excluded in parts for excluded in EXCLUDE_DIRS)


def get_file_type(ext: str) -> str:
    """获取文件类型分类"""
    return FILE_TYPE_MAP.get(ext.lower(), 'unknown')


def clean_text(text: str) -> str:
    """清理文本，移除多余空白和不可见字符"""
    if not text:
        return ''
    # 移除多个连续空白字符
    text = re.sub(r'\s+', ' ', text)
    # 移除不可见字符
    text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f]', '', text)
    return text.strip()


def hash_content(content: str) -> str:
    """计算内容hash"""
    return hashlib.md5(content.encode('utf-8', errors='ignore')).hexdigest()


# ============================================================
# 内容提取器
# ============================================================

async def extract_text_from_file(file_path: Path, ext: str) -> Optional[str]:
    """
    根据文件类型异步提取文本内容
    
    Args:
        file_path: 文件路径
        ext: 文件扩展名
        
    Returns:
        提取的文本内容，失败返回 None
    """
    try:
        if ext == '.md' or ext == '.txt':
            return await extract_text_plain(file_path)
        elif ext == '.pdf':
            return await extract_text_pdf(file_path)
        elif ext in {'.xlsx', '.xls'}:
            return await extract_text_xlsx(file_path)
        elif ext == '.docx':
            return await extract_text_docx(file_path)
        elif ext in {'.doc'}:
            return await extract_text_doc(file_path)
        elif ext == '.pptx':
            return await extract_text_pptx(file_path)
        elif ext in {'.py', '.js', '.ts'}:
            return await extract_text_plain(file_path)
        else:
            return None
    except Exception as e:
        print(f"[LocalScanner] Failed to extract {file_path}: {e}")
        return None


async def extract_text_plain(file_path: Path) -> Optional[str]:
    """提取纯文本文件内容"""
    try:
        async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = await f.read()
        return content
    except UnicodeDecodeError:
        # 尝试其他编码
        try:
            async with aiofiles.open(file_path, 'r', encoding='gbk', errors='ignore') as f:
                content = await f.read()
            return content
        except Exception:
            return None
    except Exception:
        return None


async def extract_text_pdf(file_path: Path) -> Optional[str]:
    """使用 PyMuPDF 提取 PDF 文本"""
    try:
        loop = asyncio.get_event_loop()
        text = await loop.run_in_executor(None, _extract_pdf_sync, str(file_path))
        return text
    except Exception:
        return None


def _extract_pdf_sync(file_path: str) -> Optional[str]:
    """同步 PDF 提取（在线程池中运行）"""
    try:
        doc = fitz.open(file_path)
        text_parts = []
        for page in doc:
            text_parts.append(page.get_text())
        doc.close()
        return '\n'.join(text_parts) if text_parts else None
    except Exception:
        return None


async def extract_text_xlsx(file_path: Path) -> Optional[str]:
    """使用 openpyxl 提取 Excel 文本"""
    try:
        loop = asyncio.get_event_loop()
        text = await loop.run_in_executor(None, _extract_xlsx_sync, str(file_path))
        return text
    except Exception:
        return None


def _extract_xlsx_sync(file_path: str) -> Optional[str]:
    """同步 Excel 提取（在线程池中运行）"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"[Sheet: {sheet_name}]")
            for row in sheet.iter_rows(values_only=True):
                row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return '\n'.join(text_parts) if text_parts else None
    except Exception:
        return None


async def extract_text_docx(file_path: Path) -> Optional[str]:
    """提取 DOCX 文本（使用内置 zipfile + xml）"""
    try:
        import zipfile
        from xml.etree import ElementTree as ET
        
        loop = asyncio.get_event_loop()
        text = await loop.run_in_executor(None, _extract_docx_sync, str(file_path))
        return text
    except Exception:
        return None


def _extract_docx_sync(file_path: str) -> Optional[str]:
    """同步 DOCX 提取"""
    try:
        import zipfile
        from xml.etree import ElementTree as ET
        
        text_parts = []
        with zipfile.ZipFile(file_path, 'r') as z:
            with z.open('word/document.xml') as xml_file:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                for para in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                    para_text = ''
                    for text_elem in para.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                        if text_elem.text:
                            para_text += text_elem.text
                    if para_text.strip():
                        text_parts.append(para_text)
        return '\n'.join(text_parts) if text_parts else None
    except Exception:
        return None


async def extract_text_doc(file_path: Path) -> Optional[str]:
    """提取 DOC 文本（使用 antiword 或返回简单提示）"""
    # DOC 格式较复杂，尝试使用 python-docx 或直接返回提示
    # 这里简单返回文件路径信息，实际使用中可扩展
    return None


async def extract_text_pptx(file_path: Path) -> Optional[str]:
    """提取 PPTX 文本"""
    try:
        import zipfile
        from xml.etree import ElementTree as ET
        
        loop = asyncio.get_event_loop()
        text = await loop.run_in_executor(None, _extract_pptx_sync, str(file_path))
        return text
    except Exception:
        return None


def _extract_pptx_sync(file_path: str) -> Optional[str]:
    """同步 PPTX 提取"""
    try:
        import zipfile
        from xml.etree import ElementTree as ET
        
        text_parts = []
        with zipfile.ZipFile(file_path, 'r') as z:
            slide_files = sorted([f for f in z.namelist() if f.startswith('ppt/slides/slide') and f.endswith('.xml')])
            for slide_file in slide_files:
                with z.open(slide_file) as xml_file:
                    tree = ET.parse(xml_file)
                    root = tree.getroot()
                    for elem in root.iter():
                        if elem.text and elem.text.strip():
                            text_parts.append(elem.text.strip())
        return '\n'.join(text_parts) if text_parts else None
    except Exception:
        return None


# ============================================================
# 核心扫描器
# ============================================================

async def scan_file(file_path: Path, user_id: str = '') -> Optional[Dict]:
    """
    扫描单个文件，提取元数据和内容预览
    
    Args:
        file_path: 文件路径
        user_id: 用户ID
        
    Returns:
        文件信息字典，失败返回 None
    """
    try:
        # 获取文件元数据
        stat = file_path.stat()
        
        # 跳过大于 5MB 的文件
        if stat.st_size > MAX_FILE_SIZE:
            return None
        
        ext = file_path.suffix.lower()
        
        # 检查是否支持的文件类型
        if ext not in SUPPORTED_EXTENSIONS:
            return None
        
        # 检查敏感文件名
        if is_sensitive_file(file_path):
            return None
        
        # 提取文本内容
        content = await extract_text_file(file_path, ext)
        
        # 生成内容预览（前200字纯文本）
        content_preview = ''
        if content:
            content_preview = clean_text(content)[:PREVIEW_MAX_CHARS]
        
        # 获取修改时间
        modified = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%dT%H:%M:%S+08:00')
        
        return {
            'path': str(file_path),
            'name': file_path.name,
            'ext': ext,
            'modified': modified,
            'size': stat.st_size,
            'type': get_file_type(ext),
            'content_preview': content_preview
        }
        
    except Exception as e:
        print(f"[LocalScanner] Error scanning {file_path}: {e}")
        return None


async def extract_text_file(file_path: Path, ext: str) -> Optional[str]:
    """提取文件文本内容"""
    return await extract_text_from_file(file_path, ext)


async def scan_directory(directory: Path, user_id: str = '', max_files: int = 10000) -> List[Dict]:
    """
    异步扫描目录，返回所有符合条件的文件
    
    Args:
        directory: 要扫描的目录
        user_id: 用户ID
        max_files: 最大扫描文件数（防止无限扫描）
        
    Returns:
        文件信息列表
    """
    results = []
    scanned_count = 0
    
    try:
        # 使用 os.walk 遍历目录
        for root, dirs, files in os.walk(directory):
            # 过滤排除的目录
            dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
            
            root_path = Path(root)
            
            # 检查是否在排除目录中
            if should_exclude_dir(root_path):
                continue
            
            for filename in files:
                if scanned_count >= max_files:
                    break
                    
                file_path = root_path / filename
                
                # 跳过来自排除目录的文件
                if should_exclude_dir(file_path.parent):
                    continue
                
                result = await scan_file(file_path, user_id)
                if result:
                    results.append(result)
                    scanned_count += 1
            
            if scanned_count >= max_files:
                break
                
    except PermissionError:
        print(f"[LocalScanner] Permission denied: {directory}")
    except Exception as e:
        print(f"[LocalScanner] Error scanning directory {directory}: {e}")
    
    return results


async def scan_local(
    user_id: str = '',
    scan_paths: List[str] = None,
    preferences_path: str = '~/.jobtracer/memory/preferences.json',
    output_path: str = '~/.jobtracer/scanner/results/local_files.json'
) -> Dict:
    """
    扫描本地文件的主入口函数
    
    Args:
        user_id: 用户ID
        scan_paths: 自定义扫描路径列表（优先级高于默认路径）
        preferences_path: 用户偏好配置文件路径
        output_path: 输出结果文件路径
        
    Returns:
        扫描结果字典
    """
    start_time = datetime.now()
    
    # 确定要扫描的路径（按优先级）
    paths_to_scan = []
    
    # 1. 用户自定义路径（最高优先级）
    if scan_paths:
        paths_to_scan.extend(scan_paths)
    
    # 2. 从 preferences.json 读取的路径
    pref_path = Path(preferences_path).expanduser()
    if pref_path.exists():
        try:
            with open(pref_path, 'r', encoding='utf-8') as f:
                prefs = json.load(f)
            custom_paths = prefs.get('local_scan_paths', [])
            for p in custom_paths:
                if p not in paths_to_scan:
                    paths_to_scan.append(p)
        except Exception as e:
            print(f"[LocalScanner] Failed to read preferences: {e}")
    
    # 3. 默认扫描路径
    default_paths = [
        '~/.openclaw/workspace/',
        '~/Documents/',
        '~/Desktop/',
    ]
    
    for default_path in default_paths:
        expanded = os.path.expanduser(default_path)
        if expanded not in paths_to_scan:
            paths_to_scan.append(expanded)
    
    # 去重并确保路径存在
    unique_paths = []
    for p in paths_to_scan:
        expanded = os.path.expanduser(p)
        if os.path.exists(expanded) and expanded not in unique_paths:
            unique_paths.append(expanded)
    
    # 并行扫描所有路径
    scan_tasks = []
    for path in unique_paths:
        scan_tasks.append(scan_directory(Path(path), user_id))
    
    # 执行并行扫描
    results_list = await asyncio.gather(*scan_tasks, return_exceptions=True)
    
    # 合并结果
    all_results = []
    for result in results_list:
        if isinstance(result, list):
            all_results.extend(result)
        elif isinstance(result, Exception):
            print(f"[LocalScanner] Scan task failed: {result}")
    
    # 记录扫描元数据
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    # 构建输出结果
    output = {
        'user_id': user_id,
        'scan_time': end_time.isoformat(),
        'start_time': start_time.isoformat(),
        'duration_seconds': round(duration, 2),
        'total_files': len(all_results),
        'scanned_paths': unique_paths,
        'files': all_results
    }
    
    # 确保输出目录存在
    output_file = Path(output_path).expanduser()
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 保存结果到 JSON 文件
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"[LocalScanner] Scan completed: {len(all_results)} files in {duration:.2f}s")
    print(f"[LocalScanner] Results saved to: {output_file}")
    
    return output


# ============================================================
# 便捷入口（兼容 main.py 直接调用）
# ============================================================

async def scan_local_default() -> Dict:
    """使用默认配置扫描本地文件"""
    return await scan_local()


# ============================================================
# 测试入口
# ============================================================

if __name__ == '__main__':
    async def test():
        print("[LocalScanner] Starting test scan...")
        
        # 扫描 ~/openclaw/workspace/
        result = await scan_local(
            user_id='test_user',
            scan_paths=['~/openclaw/workspace/'],
            output_path='~/openclaw-workspaces/product-solution/jobtracer/scanner/results/local_files.json'
        )
        
        print(f"\n[LocalScanner] Scan Results:")
        print(f"  Total files: {result['total_files']}")
        print(f"  Duration: {result['duration_seconds']}s")
        print(f"  Paths scanned: {result['scanned_paths']}")
        
        if result['files']:
            print(f"\n[LocalScanner] First 5 files:")
            for f in result['files'][:5]:
                print(f"  - {f['name']} ({f['type']}) | {f['ext']} | {f['size']} bytes")
        
        return result
    
    asyncio.run(test())